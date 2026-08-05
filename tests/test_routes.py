"""Testes de integração para as rotas FastAPI usando SQLite em memória."""

from __future__ import annotations

import csv
import io
import json
import logging
from typing import Any, Generator

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from sqlalchemy.exc import OperationalError
from sqlalchemy.pool import StaticPool


# ── Setup do banco de testes (SQLite) ─────────────────────────────────────────

def _criar_engine_sqlite() -> Engine:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    with engine.connect() as conn:
        conn.execute(text("PRAGMA foreign_keys = ON"))
        conn.execute(text("""
            CREATE TABLE clientes (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                nome      TEXT NOT NULL,
                email     TEXT,
                cidade    TEXT,
                criado_em TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE processos (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                numero    TEXT NOT NULL,
                status    TEXT,
                descricao TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE users (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL
            )
        """))
        conn.execute(text("""
            CREATE TABLE busunitaccess (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                userid INTEGER NOT NULL,
                perfil TEXT,
                confirmado INTEGER,
                FOREIGN KEY(userid) REFERENCES users(id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE pedidos2lawsuit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lawsuit_id INTEGER
            )
        """))
        conn.execute(text("""
            INSERT INTO clientes (nome, email, cidade, criado_em) VALUES
            ('João Silva', 'joao@ex.com', 'São Paulo', '2023-01-01 10:00:00'),
            ('Maria Souza', 'maria@ex.com', 'Rio de Janeiro', '2023-02-01 11:00:00'),
            ('Carlos Andrade', 'carlos@ex.com', 'Belo Horizonte', '2023-03-01 12:00:00')
        """))
        conn.execute(text("""
            INSERT INTO processos (numero, status, descricao) VALUES
            ('0001234-00.2023.8', 'em andamento', 'Ação de cobrança'),
            ('0002345-00.2023.8', 'concluído', 'Divórcio consensual')
        """))
        conn.execute(text("INSERT INTO users (nome) VALUES ('Ana'), ('Bruno')"))
        conn.execute(text("INSERT INTO busunitaccess (userid, perfil, confirmado) VALUES (1, 'admin', 0), (2, 'leitura', 1)"))
        conn.commit()
    return engine


# ── Fixtures e mocks ──────────────────────────────────────────────────────────

def _listar_tabelas_sqlite(engine: Engine) -> list[dict[str, Any]]:
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        )
        tabelas = []
        for row in resultado:
            nome = row[0]
            if nome.startswith("sqlite_"):
                continue
            total = conn.execute(text(f"SELECT COUNT(*) FROM `{nome}`")).scalar_one()
            tabelas.append({"nome": nome, "linhas_aprox": total, "tamanho_mb": 0.0})
    return tabelas


def _listar_colunas_sqlite(engine: Engine, nome_tabela: str) -> list[dict[str, Any]]:
    with engine.connect() as conn:
        resultado = conn.execute(text(f"PRAGMA table_info(`{nome_tabela}`)"))
        return [
            {
                "nome": row[1],
                "tipo": row[2],
                "nulo": row[3] == 0,
                "chave": "PRI" if row[5] == 1 else "",
            }
            for row in resultado
        ]


def _tabelas_validas_sqlite(engine: Engine) -> set[str]:
    return {t["nome"] for t in _listar_tabelas_sqlite(engine)}


def _colunas_validas_sqlite(engine: Engine, nome_tabela: str) -> set[str]:
    return {c["nome"] for c in _listar_colunas_sqlite(engine, nome_tabela)}


def _colunas_texto_sqlite(
    engine: Engine,
    nome_tabela: str,
    incluir_colunas_grandes: bool = False,
) -> list[str]:
    _ = incluir_colunas_grandes
    tipos_texto = {"text", "varchar", "char", "json"}
    colunas = _listar_colunas_sqlite(engine, nome_tabela)
    return [
        c["nome"]
        for c in colunas
        if c["tipo"].lower().split("(")[0] in tipos_texto
    ]


def _listar_fks_sqlite(engine: Engine, nome_tabela: str) -> list[dict[str, str]]:
    with engine.connect() as conn:
        resultado = conn.execute(text(f"PRAGMA foreign_key_list(`{nome_tabela}`)"))
        return [
            {
                "coluna": row[3],
                "tabela_referenciada": row[2],
                "coluna_referenciada": row[4],
            }
            for row in resultado
        ]


@pytest.fixture
def client(monkeypatch: pytest.MonkeyPatch) -> Generator[TestClient, None, None]:
    from src.api import main as main_module
    import src.api.routes_tables as rt_tables
    import src.api.routes_data as rt_data
    import src.api.routes_search as rt_search
    import src.api.routes_export as rt_export
    import src.api.routes_export_search as rt_export_search
    import src.api.routes_stats as rt_stats
    import src.db as db_module

    engine = _criar_engine_sqlite()

    monkeypatch.setattr(db_module, "listar_tabelas", _listar_tabelas_sqlite)
    monkeypatch.setattr(db_module, "listar_colunas", _listar_colunas_sqlite)
    monkeypatch.setattr(db_module, "tabelas_validas", _tabelas_validas_sqlite)
    monkeypatch.setattr(db_module, "colunas_validas", _colunas_validas_sqlite)
    monkeypatch.setattr(db_module, "colunas_texto", _colunas_texto_sqlite)
    monkeypatch.setattr(db_module, "listar_chaves_estrangeiras", _listar_fks_sqlite)

    for modulo in [rt_tables, rt_data, rt_search, rt_export, rt_export_search, rt_stats]:
        if hasattr(modulo, "listar_tabelas"):
            monkeypatch.setattr(modulo, "listar_tabelas", _listar_tabelas_sqlite)
        if hasattr(modulo, "listar_colunas"):
            monkeypatch.setattr(modulo, "listar_colunas", _listar_colunas_sqlite)
        if hasattr(modulo, "tabelas_validas"):
            monkeypatch.setattr(modulo, "tabelas_validas", _tabelas_validas_sqlite)
        if hasattr(modulo, "colunas_validas"):
            monkeypatch.setattr(modulo, "colunas_validas", _colunas_validas_sqlite)
        if hasattr(modulo, "colunas_texto"):
            monkeypatch.setattr(modulo, "colunas_texto", _colunas_texto_sqlite)
        if hasattr(modulo, "listar_chaves_estrangeiras"):
            monkeypatch.setattr(modulo, "listar_chaves_estrangeiras", _listar_fks_sqlite)

    from src.api.main import app
    app.state.engine = engine
    app.state.dicionarios = {"busunitaccess": {"confirmado": {"0": "Não", "1": "Sim"}}}

    with TestClient(app, raise_server_exceptions=True) as c:
        yield c


class TestRotaTabelas:
    def test_raiz_carrega_interface_com_modo_simples(self, client: TestClient) -> None:
        resp_html = client.get("/")
        assert resp_html.status_code == 200
        assert "Modo Avançado" in resp_html.text
        assert "Digite o nome do cliente ou número do processo que você procura" in resp_html.text
        assert "Buscar agora" in resp_html.text

        resp_js = client.get("/static/app.js")
        assert resp_js.status_code == 200
        assert "saidjur_modo_avancado" in resp_js.text

    def test_lista_tabelas_status_200(self, client: TestClient) -> None:
        resp = client.get("/api/tabelas")
        assert resp.status_code == 200

    def test_colunas_tabela_existente(self, client: TestClient) -> None:
        resp = client.get("/api/tabelas/clientes/colunas")
        assert resp.status_code == 200
        nomes = [c["nome"] for c in resp.json()]
        assert "nome" in nomes

    def test_fks_tabela(self, client: TestClient) -> None:
        resp = client.get("/api/tabelas/busunitaccess/fks")
        assert resp.status_code == 200
        dados = resp.json()
        assert dados[0]["coluna"] == "userid"
        assert dados[0]["tabela_referenciada"] == "users"

    def test_lista_tabelas_retry_na_primeira_falha(self, client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
        import src.api.routes_tables as rt_tables

        chamadas = {"total": 0}

        def flaky(engine: Engine) -> list[dict[str, Any]]:
            chamadas["total"] += 1
            if chamadas["total"] == 1:
                raise OperationalError("SELECT 1", {}, RuntimeError("stale connection"))
            return _listar_tabelas_sqlite(engine)

        monkeypatch.setattr(rt_tables, "listar_tabelas", flaky)

        resp = client.get("/api/tabelas")

        assert resp.status_code == 200
        assert chamadas["total"] == 2

    def test_lista_tabelas_503_faz_log(self, client: TestClient, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture) -> None:
        import src.api.routes_tables as rt_tables

        def falhar(_: Engine) -> list[dict[str, Any]]:
            raise RuntimeError("banco indisponível")

        monkeypatch.setattr(rt_tables, "listar_tabelas", falhar)

        with caplog.at_level(logging.ERROR):
            resp = client.get("/api/tabelas")

        assert resp.status_code == 503
        assert "Não foi possível conectar ao banco de dados" in resp.json()["detail"]
        assert "Falha ao listar tabelas" in caplog.text


class TestExportacaoBusca:
    def test_exporta_busca_csv_com_fk_resolvida_e_booleano_traduzido(self, client: TestClient) -> None:
        payload = {
            "dados": [
                {
                    "tabela": "busunitaccess",
                    "coluna": "userid",
                    "registros": [
                        {"id": 1, "userid": 1, "perfil": "admin", "confirmado": 0},
                        {"id": 2, "userid": 2, "perfil": "leitura", "confirmado": 1},
                    ],
                },
                {
                    "tabela": "pedidos2lawsuit",
                    "coluna": "lawsuit_id",
                    "registros": [
                        {"id": 1, "lawsuit_id": 1},
                    ],
                },
            ]
        }

        resp = client.post("/api/exportar/busca?formato=csv", json=payload)

        assert resp.status_code == 200
        texto = resp.text
        assert 'Ana (1)' in texto
        assert 'Bruno (2)' in texto
        assert 'Ação de cobrança (1)' in texto
        assert 'Não' in texto
        assert 'Sim' in texto

    def test_exporta_busca_excel_com_fk_resolvida_e_booleano_traduzido(self, client: TestClient) -> None:
        payload = {
            "dados": [
                {
                    "tabela": "busunitaccess",
                    "coluna": "userid",
                    "registros": [
                        {"id": 1, "userid": 1, "perfil": "admin", "confirmado": 0},
                    ],
                }
            ]
        }

        resp = client.post("/api/exportar/busca?formato=excel", json=payload)

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb[wb.sheetnames[0]]
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        valores = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
        linha = dict(zip(headers, valores))

        assert linha['ID do Usuário'] == 'Ana (1)'
        assert linha['Confirmado'] == 'Não'

    def test_exporta_busca_excel_simplificado_com_resumo_e_abas_de_negocio(self, client: TestClient) -> None:
        payload = {
            "termo": "Sila do Brasil",
            "dados": [
                {
                    "tabela": "lawsuits",
                    "coluna": "numero",
                    "registros": [
                        {
                            "id": 10,
                            "client_id": 1,
                            "numero": "0001234-55.2024.8.26.0100",
                            "status": "Ativo",
                            "amount": "15000.50",
                            "created_at_userid": 7,
                        }
                    ],
                },
                {
                    "tabela": "pedidos2lawsuit",
                    "coluna": "claim_text",
                    "registros": [
                        {
                            "id": 99,
                            "lawsuit_id": 10,
                            "client_id": 1,
                            "claim_text": "Pedido de tutela de urgência",
                            "instance01_amount": "5000.00",
                            "log_inserted_at": "2026-08-01 10:00:00",
                        }
                    ],
                },
                {
                    "tabela": "publicationxml_extra",
                    "coluna": "summary",
                    "registros": [
                        {
                            "publication_id": 5,
                            "client_id": 1,
                            "lawsuit_id": 10,
                            "summary": "Publicação com prazo de manifestação",
                            "pub_classification": "Urgente",
                        }
                    ],
                },
            ],
        }

        resp = client.post("/api/exportar/busca?formato=excel&modo=simplificado", json=payload)

        assert resp.status_code == 200
        workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert workbook.sheetnames[0] == "Resumo"
        assert "Processos" in workbook.sheetnames
        assert "Pedidos e Andamentos" in workbook.sheetnames
        assert "Publicações" in workbook.sheetnames

        resumo = workbook["Resumo"]
        assert resumo["A2"].value == "Busca realizada"
        assert resumo["B2"].value == "Sila do Brasil"

        processos = workbook["Processos"]
        cabecalhos = [processos.cell(row=1, column=i).value for i in range(1, processos.max_column + 1)]
        valores = [processos.cell(row=2, column=i).value for i in range(1, processos.max_column + 1)]
        linha = dict(zip(cabecalhos, valores))
        assert "Cliente" in cabecalhos
        assert "Processo" in cabecalhos
        assert "Situação" in cabecalhos
        assert "ID" not in cabecalhos
        assert linha["Processo"] == "0001234-55.2024.8.26.0100"
        assert linha["Situação"] == "Ativo"

        pedidos = workbook["Pedidos e Andamentos"]
        cabecalhos_pedidos = [pedidos.cell(row=1, column=i).value for i in range(1, pedidos.max_column + 1)]
        valores_pedidos = [pedidos.cell(row=2, column=i).value for i in range(1, pedidos.max_column + 1)]
        linha_pedidos = dict(zip(cabecalhos_pedidos, valores_pedidos))
        assert "Pedido" in cabecalhos_pedidos
        assert "Valor" in cabecalhos_pedidos
        assert "ID" not in cabecalhos_pedidos
        assert linha_pedidos["Pedido"] == "Pedido de tutela de urgência"

        # Publicações: verifica que classificação aparece como coluna própria
        publicacoes = workbook["Publicações"]
        cabecalhos_pub = [publicacoes.cell(row=1, column=i).value for i in range(1, publicacoes.max_column + 1)]
        assert "Classificação" in cabecalhos_pub
        assert "Resumo" in cabecalhos_pub

    def test_exporta_busca_excel_simplificado_com_termos_de_busca(self, client: TestClient) -> None:
        """Aba 'Termos de Busca' deve aparecer quando há dados de client_publication_search_terms."""
        payload = {
            "termo": "Sila do Brasil",
            "dados": [
                {
                    "tabela": "client_publication_search_terms",
                    "coluna": "search_term",
                    "registros": [
                        {
                            "id": 1,
                            "client_id": 5,
                            "search_term": "Sila do Brasil",
                            "created_at": "2024-03-10 08:00:00",
                        },
                        {
                            "id": 2,
                            "client_id": 5,
                            "search_term": "Sila Brasil LTDA",
                            "created_at": "2024-04-01 09:30:00",
                        },
                    ],
                },
            ],
        }
        resp = client.post("/api/exportar/busca?formato=excel&modo=simplificado", json=payload)
        assert resp.status_code == 200
        workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "Termos de Busca" in workbook.sheetnames

        termos = workbook["Termos de Busca"]
        cabecalhos = [termos.cell(row=1, column=i).value for i in range(1, termos.max_column + 1)]
        assert any("Termo" in (c or "") for c in cabecalhos)
        assert any("Cadastrado" in (c or "") for c in cabecalhos)
        # IDs técnicos não devem aparecer
        assert not any("_id" in (c or "").lower() for c in cabecalhos)

    def test_resumo_enriquecido_com_contagens_e_termos(self, client: TestClient) -> None:
        """O resumo deve conter contagens por assunto e termos de busca associados."""
        payload = {
            "termo": "Sila do Brasil",
            "dados": [
                {
                    "tabela": "lawsuits",
                    "coluna": "numero",
                    "registros": [
                        {"id": 1, "numero": "0001-01.2024.8.26.0100", "status": "Ativo"},
                        {"id": 2, "numero": "0002-01.2024.8.26.0100", "status": "Encerrado"},
                    ],
                },
                {
                    "tabela": "client_publication_search_terms",
                    "coluna": "search_term",
                    "registros": [
                        {
                            "id": 1,
                            "client_id": 5,
                            "search_term": "Sila do Brasil",
                            "created_at": "2024-03-10",
                        },
                    ],
                },
            ],
        }
        resp = client.post("/api/exportar/busca?formato=excel&modo=simplificado", json=payload)
        assert resp.status_code == 200
        workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
        resumo = workbook["Resumo"]

        # Coletar todas as linhas do resumo
        linhas_resumo = {}
        for row in resumo.iter_rows(min_row=2, values_only=True):
            if row[0]:
                linhas_resumo[row[0]] = row[1]

        assert "Total de processos" in linhas_resumo
        assert int(linhas_resumo["Total de processos"]) == 2
        assert "Termos de busca associados" in linhas_resumo
        assert "Sila do Brasil" in str(linhas_resumo["Termos de busca associados"])


    def test_linhas_status_200(self, client: TestClient) -> None:
        resp = client.get("/api/tabelas/clientes/linhas")
        assert resp.status_code == 200

    def test_total_registros(self, client: TestClient) -> None:
        resp = client.get("/api/tabelas/clientes/linhas")
        assert resp.json()["total"] == 3

    def test_filtro_contem(self, client: TestClient) -> None:
        filtros = json.dumps({"nome": {"op": "contem", "valor": "João"}})
        resp = client.get(f"/api/tabelas/clientes/linhas?filtros={filtros}")
        assert resp.json()["total"] == 1


class TestRotaBusca:
    def test_busca_retorna_lista(self, client: TestClient) -> None:
        resp = client.get("/api/busca?q=João")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_busca_streaming_retorna_eventos(self, client: TestClient) -> None:
        resp = client.get("/api/busca/stream?q=João")
        assert resp.status_code == 200
        assert "\"tipo\": \"progress\"" in resp.text
        assert "\"tipo\": \"done\"" in resp.text


class TestRotaExportar:
    def test_exportar_csv_status_200(self, client: TestClient) -> None:
        resp = client.get("/api/exportar/clientes?formato=csv")
        assert resp.status_code == 200
        assert "text/csv" in resp.headers.get("content-type", "")

    def test_exportar_csv_traduz_cabecalhos(self, client: TestClient) -> None:
        resp = client.get("/api/exportar/processos?formato=csv")

        assert resp.status_code == 200

        linhas = list(csv.reader(io.StringIO(resp.text)))
        assert linhas[0] == ["ID", "Número", "Status", "Descrição"]

    def test_exportacao_busca_excel_traduz_abas_e_colunas(self) -> None:
        from src.api.routes_export_search import _exportar_excel_busca

        conteudo = _exportar_excel_busca(
            {
                "client_publication_search_terms": [
                    {
                        "id": 1,
                        "client_id": 2,
                        "search_term": "Sila do Brasil",
                        "created_at": "2026-07-31 10:00:00",
                        "created_at_userid": 7,
                    }
                ],
                "pedidos2lawsuit": [
                    {
                        "id": 1,
                        "lawsuit_id": 99,
                        "claim_text": "Pedido de indenização",
                        "instance01_amount": 1000,
                        "instance02": "Recurso",
                        "ias": "A1",
                    }
                ],
                "publicationxml_extra": [
                    {
                        "publication_id": 10,
                        "jurify_pub_id": 20,
                        "jurify_pasta": "Jurify/2026",
                        "pub_classification": "Urgente",
                        "pub_classification_id": 5,
                        "source_api": "jurify",
                    }
                ],
            },
            {},
        )

        workbook = openpyxl.load_workbook(io.BytesIO(conteudo))

        assert "Termos de Busca do Cliente" in workbook.sheetnames
        assert "Pedidos do Processo" in workbook.sheetnames
        assert "Extras da Publicação XML" in workbook.sheetnames

        aba_termos = workbook["Termos de Busca do Cliente"]
        assert [aba_termos.cell(row=1, column=col).value for col in range(1, 6)] == [
            "ID",
            "ID do Cliente",
            "Termo de Busca",
            "Data de Criação",
            "Usuário da Criação",
        ]

        aba_pedidos = workbook["Pedidos do Processo"]
        assert [aba_pedidos.cell(row=1, column=col).value for col in range(1, 7)] == [
            "ID",
            "ID do Processo",
            "Texto do Pedido",
            "Valor na 1ª Instância",
            "2ª Instância",
            "Ias",
        ]

        aba_publicacao = workbook["Extras da Publicação XML"]
        assert [aba_publicacao.cell(row=1, column=col).value for col in range(1, 7)] == [
            "ID da Publicação",
            "ID da Publicação Jurify",
            "Pasta Jurify",
            "Classificação da Publicação",
            "ID da Classificação da Publicação",
            "API de Origem",
        ]


class TestRotasNovas:
    def test_dashboard(self, client: TestClient) -> None:
        resp = client.get("/api/dashboard")
        assert resp.status_code == 200
        dados = resp.json()
        assert "estatisticas" in dados
        assert "maiores_tabelas" in dados

    def test_dashboard_retry_na_primeira_falha(self, client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
        import src.api.routes_dashboard as rt_dashboard

        chamadas = {"total": 0}

        def flaky(engine: Engine) -> dict[str, Any]:
            chamadas["total"] += 1
            if chamadas["total"] == 1:
                raise OperationalError("SELECT 1", {}, RuntimeError("stale connection"))
            return {"estatisticas": {"total_tabelas": 1, "total_registros": 2, "tamanho_total_mb": 0}, "maiores_tabelas": []}

        monkeypatch.setattr(rt_dashboard, "dados_dashboard", flaky)

        resp = client.get("/api/dashboard")

        assert resp.status_code == 200
        assert chamadas["total"] == 2

    def test_sql_select(self, client: TestClient) -> None:
        resp = client.post("/api/sql", json={"query": "SELECT id, nome FROM clientes ORDER BY id"})
        assert resp.status_code == 200
        dados = resp.json()
        assert "linhas" in dados
        assert dados["total"] >= 1

    def test_sql_rejeita_dml(self, client: TestClient) -> None:
        resp = client.post("/api/sql", json={"query": "DELETE FROM clientes"})
        assert resp.status_code == 400

    def test_stats_coluna(self, client: TestClient) -> None:
        resp = client.get("/api/tabelas/clientes/colunas/nome/stats")
        assert resp.status_code == 200
        dados = resp.json()
        assert "nao_nulos" in dados
        assert "top_5" in dados


class TestRotaRaiz:
    def test_raiz_serve_html(self, client: TestClient) -> None:
        resp = client.get("/")
        assert resp.status_code in (200, 404)


class TestRotaTraducoes:
    def test_retorna_dicionario_de_traducoes(self, client: TestClient) -> None:
        resp = client.get("/api/traducoes/colunas")
        assert resp.status_code == 200
        dados = resp.json()
        assert isinstance(dados, dict)
        assert len(dados) > 0

    def test_contem_traducoes_essenciais(self, client: TestClient) -> None:
        resp = client.get("/api/traducoes/colunas")
        dados = resp.json()
        assert dados.get("created_at") == "Data de Criação"
        assert dados.get("updated_at") == "Data de Atualização"
        assert dados.get("name") == "Nome"
        assert dados.get("plaintiff") == "Autor"
        assert dados.get("defendant") == "Réu"
        assert dados.get("lawyer") == "Advogado"

    def test_contem_traducoes_consolidadas(self, client: TestClient) -> None:
        """Entradas únicas de cada fonte anterior devem estar presentes."""
        resp = client.get("/api/traducoes/colunas")
        dados = resp.json()
        # Da app.js
        assert dados.get("canceled") == "Cancelado"
        assert dados.get("rescheduled") == "Reagendado"
        assert dados.get("reason") == "Motivo"
        # Da traducoes_nomes_colunas.py raiz
        assert dados.get("person_id") == "ID da Pessoa"
        assert dados.get("userid") == "ID do Usuário"
        assert dados.get("total") == "Total"
        assert dados.get("file") == "Arquivo"
        # Novos termos do domínio jurídico
        assert dados.get("oab") == "OAB"
        assert dados.get("empstatus") == "Status do Funcionário"


class TestMiddlewareGlobal:
    def test_middleware_loga_excecao_nao_tratada(self, caplog: pytest.LogCaptureFixture) -> None:
        from src.api.main import app

        caminho = "/api/__teste_excecao_nao_tratada"
        if not any(getattr(route, "path", None) == caminho for route in app.router.routes):
            @app.get(caminho)
            async def _rota_explosiva() -> dict[str, str]:
                raise RuntimeError("falha inesperada")

        with TestClient(app, raise_server_exceptions=False) as client_sem_raise, caplog.at_level(logging.ERROR):
            resp = client_sem_raise.get(caminho)

        assert resp.status_code == 500
        assert resp.json()["detail"] == "Erro interno do servidor. Veja logs/app.log."
        assert "Exceção não tratada em GET /api/__teste_excecao_nao_tratada" in caplog.text
