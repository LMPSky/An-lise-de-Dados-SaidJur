"""
Rotas de Exportação de Resultados de Busca
==========================================

Módulo responsável por exportar resultados de buscas globais em múltiplos formatos:
- Excel (.xlsx) com múltiplas abas
- CSV (.csv) consolidado

Autor: Lucas Paim
Data: 2026-07-29
"""

import re
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.traducoes_colunas import (
    TRADUCOES_COLUNAS,
    traduzir_nome_coluna,
    traduzir_nome_tabela_exportacao,
)
from src.api.routes_export_search import _eh_data_zero as _eh_data_zero_export

router = APIRouter(tags=["exportar-busca"])


def traduzir_coluna(nome_coluna: str) -> str:
    """Traduz nome de coluna para português."""
    return traduzir_nome_coluna(nome_coluna)


def sanitizar_nome_arquivo(texto: str, max_len: int = 50) -> str:
    """Remove caracteres inválidos para nome de arquivo"""
    caracteres_invalidos = '<>:"/\\|?*'
    for char in caracteres_invalidos:
        texto = texto.replace(char, '_')
    return texto[:max_len].strip('_')


def criar_nome_arquivo(tipo: str, termo: str = None, tabela: str = None) -> str:
    """Cria nome de arquivo com padrão consistente"""
    data_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if tabela:
        tabela_sanitizada = sanitizar_nome_arquivo(tabela)
        base = f"busca_{tabela_sanitizada}_{data_str}"
    elif termo:
        termo_sanitizado = sanitizar_nome_arquivo(termo)
        base = f"busca_{termo_sanitizado}_{data_str}"
    else:
        base = f"resultados_busca_{data_str}"
    
    extensao = 'xlsx' if tipo == 'excel' else 'csv'
    return f"{base}.{extensao}"


def criar_workbook_excel(resultados: List[Dict]) -> io.BytesIO:
    """
    Cria workbook Excel com múltiplas abas.
    
    Args:
        resultados: Lista com {'tabela': str, 'registros': [dicts]}
    
    Returns:
        Arquivo Excel em bytes
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agrupa por tabela
    tabelas_dados = {}
    for resultado in resultados:
        tabela = resultado.get('tabela', 'Dados')
        if tabela not in tabelas_dados:
            tabelas_dados[tabela] = []
        tabelas_dados[tabela].extend(resultado.get('registros', []))
    
    # Cria abas
    for tabela, registros in tabelas_dados.items():
        if not registros:
            continue
        
        ws = wb.create_sheet(title=traduzir_nome_tabela_exportacao(tabela)[:31])
        
        # Cabeçalho
        colunas = list(registros[0].keys())
        for col_idx, col_nome in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = traduzir_coluna(col_nome)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados
        for row_idx, registro in enumerate(registros, 2):
            for col_idx, col_nome in enumerate(colunas, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                valor = registro.get(col_nome, '')
                # Normaliza datas zeradas do MySQL para célula vazia
                if _eh_data_zero_export(valor):
                    valor = ''
                cell.value = valor if valor is not None else ''
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
        
        # Ajusta largura
        for col_idx, col_nome in enumerate(colunas, 1):
            max_len = max(
                len(str(registro.get(col_nome, ''))) 
                for registro in registros
            ) + 2
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len, 50)
    
    arquivo = io.BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo


def criar_csv_consolidado(resultados: List[Dict]) -> io.BytesIO:
    """
    Cria CSV consolidado com coluna de tabela.
    
    Args:
        resultados: Lista com {'tabela': str, 'registros': [dicts]}
    
    Returns:
        Arquivo CSV em bytes
    """
    arquivo = io.BytesIO()
    
    # Coleta colunas únicas
    todas_colunas = set()
    for resultado in resultados:
        for registro in resultado.get('registros', []):
            todas_colunas.update(registro.keys())
    
    colunas_ordenadas = ['tabela'] + sorted(todas_colunas)
    
    # Cria CSV
    texto = io.StringIO()
    writer = csv.writer(texto)
    writer.writerow([traduzir_nome_coluna(coluna) for coluna in colunas_ordenadas])
    
    for resultado in resultados:
        tabela = resultado.get('tabela', 'Dados')
        for registro in resultado.get('registros', []):
            linha = {'tabela': tabela}
            linha.update(registro)
            writer.writerow([linha.get(coluna, '') for coluna in colunas_ordenadas])
    
    # Salva com BOM UTF-8
    conteudo = texto.getvalue()
    arquivo.write('\ufeff'.encode('utf-8'))
    arquivo.write(conteudo.encode('utf-8'))
    arquivo.seek(0)
    
    return arquivo


# ==================
# ROTAS DE EXPORTAÇÃO
# ==================

@router.post("/exportar/busca")
async def exportar_busca_excel(termo: str, resultados: List[Dict[str, Any]]):
    """
    Exporta resultados de busca em Excel com múltiplas abas.
    
    **Estrutura esperada:**
    ```json
    {
      "termo": "ATIVO",
      "resultados": [
        {
          "tabela": "processos",
          "registros": [
            {"id": 1, "status": "ATIVO", "numero": "0001234-56..."}
          ]
        }
      ]
    }
    ```
    """
    try:
        arquivo = criar_workbook_excel(resultados)
        nome = criar_nome_arquivo('excel', termo)
        
        return StreamingResponse(
            iter([arquivo.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar Excel: {str(e)}")


@router.post("/exportar/busca/csv")
async def exportar_busca_csv(termo: str, resultados: List[Dict[str, Any]]):
    """
    Exporta resultados de busca em CSV consolidado.
    
    Inclui coluna 'tabela' para identificar origem dos dados.
    """
    try:
        arquivo = criar_csv_consolidado(resultados)
        nome = criar_nome_arquivo('csv', termo)
        
        return StreamingResponse(
            iter([arquivo.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={nome}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar CSV: {str(e)}")


@router.post("/exportar/tabela")
async def exportar_tabela_excel(tabela: str, registros: List[Dict[str, Any]], termo: str = None):
    """Exporta dados de uma tabela específica em Excel"""
    try:
        if not registros:
            raise HTTPException(status_code=400, detail="Nenhum registro para exportar")
        
        resultado_formatado = [{'tabela': tabela, 'registros': registros}]
        arquivo = criar_workbook_excel(resultado_formatado)
        nome = criar_nome_arquivo('excel', termo, tabela)
        
        return StreamingResponse(
            iter([arquivo.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar Excel: {str(e)}")


@router.post("/exportar/tabela/csv")
async def exportar_tabela_csv(tabela: str, registros: List[Dict[str, Any]], termo: str = None):
    """Exporta dados de uma tabela específica em CSV"""
    try:
        if not registros:
            raise HTTPException(status_code=400, detail="Nenhum registro para exportar")
        
        resultado_formatado = [{'tabela': tabela, 'registros': registros}]
        arquivo = criar_csv_consolidado(resultado_formatado)
        nome = criar_nome_arquivo('csv', termo, tabela)
        
        return StreamingResponse(
            iter([arquivo.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={nome}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar CSV: {str(e)}")


@router.get("/exportar/info")
async def info_exportacao():
    """Informações sobre o módulo de exportação (debug)"""
    return {
        "versao": "1.0.0",
        "rotas": [
            "POST /api/exportar/busca",
            "POST /api/exportar/busca/csv",
            "POST /api/exportar/tabela",
            "POST /api/exportar/tabela/csv"
        ],
        "formatos": ["excel (.xlsx)", "csv"],
        "codificacao": "UTF-8",
        "traducao": "Colunas automaticamente traduzidas para português"
    }
