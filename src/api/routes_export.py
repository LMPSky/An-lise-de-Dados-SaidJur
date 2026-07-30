"""Rotas para exportar dados em CSV e Excel com traduções."""

from __future__ import annotations

import io
import logging
from typing import Any

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
import csv
import json
from sqlalchemy import select, and_, text, MetaData, Table
from sqlalchemy.orm import Session

logger = logging.getLogger("saidjur.routes_export")

router = APIRouter(tags=["Exportação"])


def _carregar_dicionarios(app_state) -> dict:
    """Carrega os dicionários de tradução."""
    if hasattr(app_state, 'dicionarios'):
        return app_state.dicionarios
    return {}


def _traduzir_valor(
    valor: Any,
    tabela: str,
    coluna: str,
    dicionarios: dict,
) -> str:
    """
    Traduz um valor usando o dicionário.
    Se não encontrar tradução, retorna o valor original.
    """
    if valor is None:
        return ""
    
    valor_str = str(valor)
    
    # Procurar no dicionário
    if tabela in dicionarios:
        tabela_dict = dicionarios[tabela]
        if coluna in tabela_dict:
            coluna_dict = tabela_dict[coluna]
            if valor_str in coluna_dict:
                return coluna_dict[valor_str]
    
    # Se não encontrou, retorna original
    return valor_str


def _construir_query_com_filtros(
    engine,
    tabela_nome: str,
    filtros: dict[str, Any] | None = None,
    ordem_coluna: str | None = None,
    direcao: str = "asc",
) -> tuple:
    """Constrói query com filtros e ordenação."""
    try:
        metadata = MetaData()
        table = Table(tabela_nome, metadata, autoload_with=engine)
        
        stmt = select(table)
        
        # Aplicar filtros
        if filtros:
            condicoes = []
            for coluna, filtro_info in filtros.items():
                col = table.columns.get(coluna)
                if col is None:
                    continue
                
                op = filtro_info.get("op", "contem")
                valor = filtro_info.get("valor", "")
                
                if op == "igual":
                    condicoes.append(col == valor)
                elif op == "contem":
                    condicoes.append(col.ilike(f"%{valor}%"))
                elif op == "comeca":
                    condicoes.append(col.ilike(f"{valor}%"))
                elif op == "termina":
                    condicoes.append(col.ilike(f"%{valor}"))
            
            if condicoes:
                stmt = stmt.where(and_(*condicoes))
        
        # Ordenação
        if ordem_coluna:
            col = table.columns.get(ordem_coluna)
            if col is not None:
                if direcao.lower() == "desc":
                    stmt = stmt.order_by(col.desc())
                else:
                    stmt = stmt.order_by(col.asc())
        
        return table, stmt
    
    except Exception as e:
        logger.error(f"Erro ao construir query para {tabela_nome}: {e}")
        raise HTTPException(status_code=400, detail=str(e))


def _serializar_valor(valor: Any) -> str:
    """Converte valor para string exportável."""
    if valor is None:
        return ""
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False)
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    return str(valor)


@router.get("/exportar/{tabela}")
async def exportar_tabela(
    tabela: str,
    request: Request,
    formato: str = "csv",
    filtros: str | None = None,
) -> StreamingResponse:
    """
    Exporta dados de uma tabela em CSV ou Excel com traduções.
    
    **Parâmetros:**
    - `tabela`: nome da tabela
    - `formato`: 'csv' ou 'excel'
    - `filtros`: JSON com filtros (opcional)
    
    **Exemplo:**
    ```
    GET /api/exportar/clientes?formato=csv&filtros={"status":{"op":"igual","valor":"ATIVO"}}
    ```
    """
    
    engine = request.app.state.engine
    dicionarios = _carregar_dicionarios(request.app.state)
    
    # Validar formato
    if formato.lower() not in ("csv", "excel", "xlsx"):
        raise HTTPException(status_code=400, detail="Formato deve ser 'csv' ou 'excel'")
    
    # Parsear filtros
    filtros_dict = {}
    if filtros:
        try:
            filtros_dict = json.loads(filtros)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Filtros JSON inválido")
    
    try:
        table, stmt = _construir_query_com_filtros(
            engine=engine,
            tabela_nome=tabela,
            filtros=filtros_dict,
        )
        
        with Session(engine) as session:
            resultado = session.execute(stmt).fetchall()
            colunas = [col.name for col in table.columns]
            
            if formato.lower() == "csv":
                # Exportar como CSV
                output = io.StringIO()
                writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_ALL)
                
                # Cabeçalho
                writer.writerow(colunas)
                
                # Dados com tradução
                for row in resultado:
                    row_dict = row._mapping
                    linha_traduzida = []
                    for col_nome in colunas:
                        valor_original = row_dict.get(col_nome)
                        valor_traduzido = _traduzir_valor(
                            valor_original,
                            tabela,
                            col_nome,
                            dicionarios,
                        )
                        linha_traduzida.append(valor_traduzido)
                    writer.writerow(linha_traduzida)
                
                # Preparar resposta
                output.seek(0)
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f"attachment; filename={tabela}.csv"},
                )
            
            else:  # excel/xlsx
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    raise HTTPException(
                        status_code=500,
                        detail="Biblioteca openpyxl não instalada. Execute: pip install openpyxl"
                    )
                
                # Criar workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = tabela[:31]  # Excel limita nome de sheet a 31 caracteres
                
                # Cabeçalho com estilo
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, col_nome in enumerate(colunas, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = col_nome
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Dados com tradução
                for row_idx, row in enumerate(resultado, 2):
                    row_dict = row._mapping
                    for col_idx, col_nome in enumerate(colunas, 1):
                        valor_original = row_dict.get(col_nome)
                        valor_traduzido = _traduzir_valor(
                            valor_original,
                            tabela,
                            col_nome,
                            dicionarios,
                        )
                        
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = valor_traduzido
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Ajustar largura das colunas
                for col_idx, col_nome in enumerate(colunas, 1):
                    max_length = len(col_nome)
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
                
                # Salvar em memória
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={tabela}.xlsx"},
                )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Erro ao exportar {tabela}")
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")