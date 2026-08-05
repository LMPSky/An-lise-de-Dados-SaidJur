"""Rotas para exportar resultados de busca em Excel e CSV."""

from __future__ import annotations

import io
import json
import logging
import re
from collections import defaultdict
from typing import Any

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
import csv

from src.db import fks_inferidas, listar_chaves_estrangeiras, resolver_labels
from src.traducoes_colunas import (
    traduzir_nome_coluna,
    traduzir_nome_tabela_exportacao,
)

logger = logging.getLogger("saidjur.export_search")

router = APIRouter(tags=["Exportação"])

_TABELAS_VISUALIZACAO_SIMPLES = {
    "lawsuits": "Processos",
    "publicationxml": "Publicações",
    "publicationxml_extra": "Publicações",
    "hearingcontrol": "Audiências",
    "pedidos2lawsuit": "Pedidos e Andamentos",
    "clients": "Clientes",
    "persons": "Partes",
}

_ROTULOS_SIMPLES = {
    "clients": {
        "name": "Cliente",
        "nome": "Cliente",
    },
    "persons": {
        "name": "Parte",
        "nome": "Parte",
    },
    "lawsuits": {
        "numero": "Processo",
        "lawsuitnumber": "Processo",
        "status": "Situação",
        "instance": "Instância",
        "amount": "Valor",
        "value": "Valor",
        "description": "Descrição",
        "subject": "Assunto",
        "type": "Tipo",
    },
    "publicationxml": {
        "date": "Data",
        "publication": "Publicação",
        "summary": "Resumo",
        "content": "Texto",
        "type": "Tipo",
        "status": "Situação",
    },
    "publicationxml_extra": {
        "classification": "Classificação",
        "summary": "Resumo",
        "content": "Texto",
    },
    "hearingcontrol": {
        "date": "Data",
        "type": "Tipo de Audiência",
        "status": "Situação",
        "room": "Local",
        "description": "Descrição",
    },
    "pedidos2lawsuit": {
        "claim": "Pedido",
        "request": "Pedido",
        "status": "Situação",
        "date": "Data",
        "amount": "Valor",
        "instance": "Instância",
        "progress": "Andamento",
        "text": "Texto",
    },
}

_COLUNAS_TECNICAS_REGEX = (
    r"(^id$|_id$|^fk_|^id_|created_at|updated_at|deleted_at|inserted_at|"
    r"created_by|updated_by|userid|user_id|log_|config|setting|token|hash|password|"
    r"checksum|uuid|guid|version|sort_order|ordem|ativo$|enabled$)"
)


def _traduzir_valor_coluna(
    valor: Any,
    tabela: str,
    coluna: str,
    dicionarios: dict[str, Any] | None = None,
) -> Any:
    """Traduz um valor usando o dicionário já carregado na aplicação."""
    if valor is None:
        return None

    dicionarios = dicionarios or {}
    traducao = dicionarios.get(tabela, {}).get(coluna, {}).get(str(valor))
    return traducao if traducao is not None else valor


def _serializar_valor(valor: Any) -> str:
    """Converte valor para string exportável."""
    if valor is None:
        return ""
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False)
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    return str(valor)


def _eh_coluna_tecnica(nome_coluna: str) -> bool:
    """Indica se a coluna é técnica demais para o relatório simplificado."""
    return bool(re.search(_COLUNAS_TECNICAS_REGEX, nome_coluna.lower()))


def _eh_texto_vazio(valor: Any) -> bool:
    """Retorna True quando o valor é vazio para a visão simplificada."""
    return valor is None or str(valor).strip() == ""


def _rotulo_simples_para_coluna(tabela: str, coluna: str) -> str | None:
    """Tenta gerar um cabeçalho amigável por heurística."""
    coluna_lower = coluna.lower()
    for trecho, rotulo in _ROTULOS_SIMPLES.get(tabela, {}).items():
        if trecho in coluna_lower:
            return rotulo
    return None


def _coluna_valor_mais_relevante(registro: dict[str, Any], candidatos: list[str]) -> Any:
    """Retorna o primeiro valor não vazio encontrado entre colunas candidatas."""
    for nome in candidatos:
        if nome in registro and not _eh_texto_vazio(registro[nome]):
            return registro[nome]
    return None


def _montar_registro_simplificado(tabela: str, registro: dict[str, Any]) -> dict[str, Any]:
    """Converte um registro técnico em um registro amigável para leigos."""
    if tabela == "lawsuits":
        simplificado = {
            "Cliente": _coluna_valor_mais_relevante(registro, ["client_id", "client_name", "cliente", "nome_cliente"]),
            "Processo": _coluna_valor_mais_relevante(registro, ["numero", "lawsuitnumber", "cnj", "number"]),
            "Parte": _coluna_valor_mais_relevante(registro, ["person_id", "person_name", "parte", "nome_parte"]),
            "Situação": _coluna_valor_mais_relevante(registro, ["status", "situation", "phase"]),
            "Valor": _coluna_valor_mais_relevante(registro, ["amount", "value", "valor_causa", "instance01_amount", "total_amount"]),
            "Descrição": _coluna_valor_mais_relevante(registro, ["description", "subject", "resumo", "summary"]),
        }
    elif tabela in {"publicationxml", "publicationxml_extra"}:
        simplificado = {
            "Cliente": _coluna_valor_mais_relevante(registro, ["client_id", "client_name", "cliente"]),
            "Processo": _coluna_valor_mais_relevante(registro, ["lawsuit_id", "numero", "lawsuitnumber", "processo"]),
            "Data": _coluna_valor_mais_relevante(registro, ["publication_date", "date", "created_at"]),
            "Situação": _coluna_valor_mais_relevante(registro, ["status", "pub_classification", "classification"]),
            "Resumo": _coluna_valor_mais_relevante(registro, ["summary", "publication", "content", "texto"]),
        }
    elif tabela == "hearingcontrol":
        simplificado = {
            "Cliente": _coluna_valor_mais_relevante(registro, ["client_id", "client_name", "cliente"]),
            "Processo": _coluna_valor_mais_relevante(registro, ["lawsuit_id", "numero", "lawsuitnumber"]),
            "Parte": _coluna_valor_mais_relevante(registro, ["person_id", "person_name"]),
            "Data": _coluna_valor_mais_relevante(registro, ["hearing_date", "date", "scheduled_at"]),
            "Tipo de Audiência": _coluna_valor_mais_relevante(registro, ["hearing_type_id", "type", "hearing_type"]),
            "Situação": _coluna_valor_mais_relevante(registro, ["status", "situation"]),
        }
    elif tabela == "pedidos2lawsuit":
        simplificado = {
            "Cliente": _coluna_valor_mais_relevante(registro, ["client_id", "client_name", "cliente"]),
            "Processo": _coluna_valor_mais_relevante(registro, ["lawsuit_id", "numero", "lawsuitnumber"]),
            "Pedido": _coluna_valor_mais_relevante(registro, ["claim_text", "request_text", "pedido", "description"]),
            "Andamento": _coluna_valor_mais_relevante(registro, ["progress_text", "status", "instance02", "instance01"]),
            "Valor": _coluna_valor_mais_relevante(registro, ["instance01_amount", "amount", "value"]),
        }
    else:
        simplificado = {}

    if simplificado:
        return {chave: valor for chave, valor in simplificado.items() if not _eh_texto_vazio(valor)}

    fallback: dict[str, Any] = {}
    for coluna, valor in registro.items():
        if _eh_texto_vazio(valor) or _eh_coluna_tecnica(coluna):
            continue
        rotulo = _rotulo_simples_para_coluna(tabela, coluna) or traduzir_nome_coluna(coluna)
        if rotulo not in fallback:
            fallback[rotulo] = valor
    return fallback


def montar_relatorio_simplificado(
    dados_por_tabela: dict[str, list[dict]],
    termo_busca: str = "",
) -> dict[str, list[dict]]:
    """Monta visões consolidadas e amigáveis por assunto de negócio."""
    consolidado: dict[str, list[dict]] = defaultdict(list)
    tabelas_cobertas: set[str] = set()
    total_registros = 0

    for tabela, registros in dados_por_tabela.items():
        assunto = _TABELAS_VISUALIZACAO_SIMPLES.get(tabela)
        if not assunto:
            continue
        tabelas_cobertas.add(tabela)
        for registro in registros:
            linha = _montar_registro_simplificado(tabela, registro)
            if linha:
                consolidado[assunto].append(linha)
                total_registros += 1

    nao_cobertas = sorted(t for t in dados_por_tabela.keys() if t not in tabelas_cobertas)
    resumo = [
        {"Informação": "Busca realizada", "Detalhe": termo_busca or "Busca sem termo informado"},
        {"Informação": "O que este relatório mostra", "Detalhe": "Informações principais sobre processos, publicações, audiências e pedidos."},
        {"Informação": "Total de registros simplificados", "Detalhe": total_registros},
        {"Informação": "Assuntos incluídos", "Detalhe": ", ".join(consolidado.keys()) or "Nenhum assunto compatível encontrado"},
    ]
    if nao_cobertas:
        resumo.append(
            {
                "Informação": "Tabelas ainda não simplificadas",
                "Detalhe": ", ".join(nao_cobertas),
            }
        )

    return {"Resumo": resumo, **consolidado}


def _formatar_label_fk(label: str, valor_original: Any) -> str:
    """Formata label de FK mantendo o ID visível para rastreabilidade."""
    valor_str = _serializar_valor(valor_original)
    return f"{label} ({valor_str})" if label != valor_str else valor_str


def _mapear_fks_para_exportacao(engine, tabela: str) -> dict[str, str]:
    """Retorna mapa coluna FK → tabela referenciada para exportação."""
    fks = listar_chaves_estrangeiras(engine, tabela) + fks_inferidas(engine, tabela)
    mapeamento: dict[str, str] = {}
    for fk in fks:
        coluna = fk.get("coluna")
        tabela_ref = fk.get("tabela_referenciada")
        if coluna and tabela_ref and coluna not in mapeamento:
            mapeamento[coluna] = tabela_ref
    return mapeamento


def _resolver_fks_dados_busca(engine, dados_por_tabela: dict[str, list[dict]]) -> dict[tuple[str, str], dict[str, str]]:
    """Resolve em lote os labels de FKs presentes nos dados exportados."""
    pedidos: dict[str, set[str]] = defaultdict(set)
    mapeamentos: dict[tuple[str, str], str] = {}

    for tabela, registros in dados_por_tabela.items():
        mapa_fks = _mapear_fks_para_exportacao(engine, tabela)
        if not mapa_fks:
            continue

        for coluna, tabela_ref in mapa_fks.items():
            for registro in registros:
                valor = registro.get(coluna)
                if valor in (None, ""):
                    continue
                valor_str = str(valor)
                if not valor_str.strip():
                    continue
                pedidos[tabela_ref].add(valor_str)
                mapeamentos[(tabela, coluna)] = tabela_ref

    resolucoes = [
        {"tabela": tabela_ref, "coluna_chave": "id", "ids": sorted(ids)}
        for tabela_ref, ids in pedidos.items()
        if ids
    ]
    if not resolucoes:
        return {}

    labels = resolver_labels(engine, resolucoes)
    return {chave: labels.get(tabela_ref, {}) for chave, tabela_ref in mapeamentos.items()}


def _normalizar_dados_para_exportacao(
    engine,
    dados_por_tabela: dict[str, list[dict]],
    dicionarios: dict[str, Any] | None = None,
) -> dict[str, list[dict]]:
    """Aplica traduções de dicionário e resolução de FK aos dados da exportação."""
    labels_fk = _resolver_fks_dados_busca(engine, dados_por_tabela)
    normalizados: dict[str, list[dict]] = {}

    for tabela, registros in dados_por_tabela.items():
        registros_normalizados: list[dict] = []
        for registro in registros:
            novo_registro: dict[str, Any] = {}
            for coluna, valor in registro.items():
                label_fk = labels_fk.get((tabela, coluna), {}).get(str(valor))
                if label_fk:
                    novo_registro[coluna] = _formatar_label_fk(label_fk, valor)
                    continue

                traducao = _traduzir_valor_coluna(valor, tabela, coluna, dicionarios)
                novo_registro[coluna] = traducao
            registros_normalizados.append(novo_registro)
        normalizados[tabela] = registros_normalizados

    return normalizados


def _extrair_dados_busca(dados_json: str) -> dict[str, list[dict]]:
    """
    Extrai dados de busca do JSON e organiza por tabela.
    
    Formato esperado:
    [
        {
            "tabela": "hearings",
            "coluna": "description",
            "registros": [{"id": 1, "description": "..."},...]
        },
        ...
    ]
    """
    try:
        dados = json.loads(dados_json)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Dados de busca JSON inválido")
    
    # Reorganizar por tabela
    por_tabela: dict[str, list[dict]] = {}
    
    if isinstance(dados, list):
        for grupo in dados:
            if isinstance(grupo, dict) and "tabela" in grupo and "registros" in grupo:
                tabela = grupo["tabela"]
                registros = grupo["registros"]
                
                if tabela not in por_tabela:
                    por_tabela[tabela] = []
                
                # Adicionar registros, evitando duplicatas (por ID se existir)
                ids_existentes = {
                    r.get("id") for r in por_tabela[tabela] if r.get("id")
                }
                
                for registro in registros:
                    if registro.get("id"):
                        if registro["id"] not in ids_existentes:
                            por_tabela[tabela].append(registro)
                            ids_existentes.add(registro["id"])
                    else:
                        por_tabela[tabela].append(registro)
    
    return por_tabela


def _exportar_csv_busca(
    dados_por_tabela: dict[str, list[dict]],
    dicionarios: dict[str, Any] | None = None,
) -> str:
    """Exporta resultados de busca como CSV com uma tabela por linha."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_ALL)
    
    # Cabeçalho com tabelas
    tabelas = list(dados_por_tabela.keys())
    writer.writerow(
        ["Tabela"] + [traduzir_nome_tabela_exportacao(tabela) for tabela in tabelas]
    )
    
    # Para cada linha, mostrar um registro de cada tabela
    max_registros = max(len(regs) for regs in dados_por_tabela.values()) if dados_por_tabela else 0
    
    for idx in range(max_registros):
        linha = []
        for tabela in tabelas:
            registros = dados_por_tabela[tabela]
            if idx < len(registros):
                registro = registros[idx]
                # Serializar o registro como JSON compacto
                linha.append(json.dumps(registro, ensure_ascii=False))
            else:
                linha.append("")
        
        writer.writerow([f"Registro {idx + 1}"] + linha)
    
    return output.getvalue()


def _exportar_excel_busca(
    dados_por_tabela: dict[str, list[dict]],
    dicionarios: dict[str, Any] | None = None,
) -> bytes:
    """Exporta resultados de busca como Excel com uma aba por tabela."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca openpyxl não instalada. Execute: pip install openpyxl"
        )
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove a planilha vazia padrão
    
    for tabela, registros in dados_por_tabela.items():
        if not registros:
            continue
        
        # Criar nova aba (limitado a 31 caracteres no Excel)
        nome_aba = traduzir_nome_tabela_exportacao(tabela)[:31]
        ws = wb.create_sheet(title=nome_aba)
        
        # Extrair colunas do primeiro registro
        if registros:
            colunas = list(registros[0].keys())
            
            # Cabeçalho com estilo
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_nome in enumerate(colunas, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = traduzir_nome_coluna(col_nome)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Dados
            for row_idx, registro in enumerate(registros, 2):
                for col_idx, col_nome in enumerate(colunas, 1):
                    valor = registro.get(col_nome)
                    valor_serializado = _serializar_valor(valor)
                    
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = valor_serializado
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Ajustar largura das colunas
            for col_idx, col_nome in enumerate(colunas, 1):
                max_length = len(traduzir_nome_coluna(col_nome))
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
    
    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def _exportar_excel_busca_simplificada(
    dados_por_tabela: dict[str, list[dict]],
    termo_busca: str = "",
) -> bytes:
    """Exporta resultados em formato simplificado por assunto de negócio."""
    visoes = montar_relatorio_simplificado(dados_por_tabela, termo_busca=termo_busca)
    return _exportar_excel_busca(visoes)


@router.post("/exportar/busca")
async def exportar_resultado_busca(
    request: Request,
    formato: str = Query("excel", pattern="^(csv|excel)$", description="Formato: 'csv' ou 'excel'"),
    tabela: str | None = Query(None, description="Exportar apenas uma tabela específica (opcional)"),
    modo: str = Query("tecnico", pattern="^(tecnico|simplificado)$", description="Modo da exportação"),
) -> StreamingResponse:
    """
    Exporta resultados de busca em Excel (múltiplas abas) ou CSV.
    
    **Parâmetros:**
    - `formato`: 'csv' ou 'excel'
    - `tabela`: (opcional) se informado, exporta apenas essa tabela
    
    **Body (JSON):**
    ```json
    {
        "dados": [
            {
                "tabela": "hearings",
                "coluna": "description",
                "registros": [{"id": 1, "description": "..."}, ...]
            },
            ...
        ]
    }
    ```
    
    **Resposta:** Arquivo baixável (.xlsx ou .csv)
    """
    
    try:
        body = await request.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Body JSON inválido: {str(e)}")
    
    dados_json = json.dumps(body.get("dados", []))
    termo_busca = body.get("termo", "")
    engine = request.app.state.engine
    dicionarios = getattr(request.app.state, "dicionarios", {})

    # Extrair dados por tabela
    dados_por_tabela = _extrair_dados_busca(dados_json)
    
    # Filtrar por tabela se especificada
    if tabela:
        if tabela not in dados_por_tabela:
            raise HTTPException(status_code=404, detail=f"Tabela '{tabela}' não encontrada nos resultados")
        dados_por_tabela = {tabela: dados_por_tabela[tabela]}
    
    if not dados_por_tabela:
        raise HTTPException(status_code=400, detail="Nenhum dado para exportar")

    dados_por_tabela = _normalizar_dados_para_exportacao(engine, dados_por_tabela, dicionarios)

    try:
        if formato.lower() == "csv":
            # Exportar como CSV
            conteudo = _exportar_csv_busca(dados_por_tabela)
            nome_arquivo = f"busca_saidjur.csv"
            
            return StreamingResponse(
                iter([conteudo]),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
            )
        
        else:  # excel
            # Exportar como Excel
            if modo == "simplificado":
                conteudo = _exportar_excel_busca_simplificada(dados_por_tabela, termo_busca=termo_busca)
                nome_arquivo = "relatorio_simplificado_saidjur.xlsx"
            else:
                conteudo = _exportar_excel_busca(dados_por_tabela)
                nome_arquivo = "busca_saidjur.xlsx"
            
            return StreamingResponse(
                iter([conteudo]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Erro ao exportar resultado de busca")
        raise HTTPException(status_code=500, detail=f"Erro ao exportar: {str(e)}")
